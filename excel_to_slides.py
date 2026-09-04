import sys
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

# --- НАСТРОЙКИ ПОД ТВ (Full HD) ---
TARGET_WIDTH = 1920
TARGET_HEIGHT = 1080

# Размер ячейки ДО масштабирования (виртуальный)
CELL_W_LOGICAL = 150
CELL_H_LOGICAL = 36

def get_font(size=18):
    for f in ["arial.ttf", "DejaVuSans.ttf", "LiberationSans-Regular.ttf", "sans-serif.ttf"]:
        try:
            return ImageFont.truetype(f, size)
        except IOError:
            continue
    return ImageFont.load_default()

def excel_to_slides_scaled(xlsx_path, output_prefix="slide"):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    font = get_font(16)  # чуть меньше, потому что масштаб сожмёт

    global_page_num = 1

    print(f"Processing workbook: {xlsx_path}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column
        max_row = ws.max_row

        if max_col == 1 and max_row == 1 and ws.cell(1, 1).value is None:
            print(f"  Skip empty sheet: {sheet_name}")
            continue

        print(f"  Sheet: {sheet_name} ({max_col} cols x {max_row} rows)")

        # Вычисляем коэффициент масштабирования, чтобы вся ширина таблицы влезла в TARGET_WIDTH
        table_width_logical = max_col * CELL_W_LOGICAL
        scale = TARGET_WIDTH / table_width_logical if table_width_logical > 0 else 1.0
        # Ограничим минимальный масштаб, чтобы не было совсем мелко
        scale = max(scale, 0.4)

        CELL_W = CELL_W_LOGICAL * scale
        CELL_H = CELL_H_LOGICAL * scale

        cols_per_page = int(TARGET_WIDTH // CELL_W)  # сколько влезет на один слайд по ширине
        rows_per_page = int(TARGET_HEIGHT // CELL_H)  # по высоте

        start_row = 1
        sheet_slides = 0

        while start_row <= max_row:
            end_row = min(start_row + rows_per_page - 1, max_row)
            start_col = 1
            end_col = min(start_col + cols_per_page - 1, max_col)

            # Ширина и высота картинки для этого слайда (всегда TARGET_WIDTH, высота может быть меньше)
            img_w = TARGET_WIDTH
            img_h = int((end_row - start_row + 1) * CELL_H)

            img = Image.new('RGB', (img_w, img_h), color='white')
            draw = ImageDraw.Draw(img)

            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    cell = ws.cell(row=r, column=c)
                    val = str(cell.value) if cell.value is not None else ""

                    x = int((c - start_col) * CELL_W)
                    y = int((r - start_row) * CELL_H)

                    # Рисуем границы ячеек
                    draw.rectangle(
                        [x, y, x + int(CELL_W) - 1, y + int(CELL_H) - 1],
                        outline="#cccccc"
                    )
                    # Текст
                    draw.text((x + 4, y + 2), val, fill="black", font=font)

            slide_name = f"{output_prefix}_{global_page_num:03d}.jpg"
            img.save(slide_name, "JPEG", quality=95)
            print(f"    -> Saved: {slide_name} (sheet: {sheet_name})")

            global_page_num += 1
            sheet_slides += 1
            start_row = end_row + 1

        print(f"  Sheet {sheet_name}: {sheet_slides} slide(s)")

    print("Done.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_slides_scaled.py file.xlsx")
        sys.exit(1)

    try:
        excel_to_slides_scaled(sys.argv[1])
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
