import os
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tqdm import tqdm
from colorama import Fore, Style, init

# Инициализация colorama (для цветного вывода в консоли Windows)
init(autoreset=True)

# --- НАСТРОЙКИ ---
XML_FOLDER = "XML-NEW"            # папка с XML-файлами
XLSX_FILE = "list-points.xlsx"    # целевой Excel-файл
HTML_FILE = "list-points.html"    # целевой HTML-файл
START_ROW = 11                    # строка, куда вставляем первую новую запись (нумерация с 1)
MARKER = "<!-- INSERT_DIMON_BLOCK_HERE -->"  # маркер для вставки блока

def parse_xml(file_path: Path) -> dict | None:
    """
    Парсит XML и возвращает словарь с нужными полями.
    Возвращает None, если файл невалиден или структура не совпадает.
    """
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()

        # Проверка, что корневой элемент — inventory
        if root.tag != "inventory":
            return None

        def get_text(tag: str) -> str:
            elem = root.find(tag)
            return elem.text.strip() if elem is not None and elem.text else ""

        data = {
            "id_date": get_text("id_date"),
            "id_mb": get_text("id_mb"),
            "id_cpu": get_text("id_cpu"),
            "id_mem": get_text("id_mem"),
            "id_hdd": get_text("id_hdd"),
            "id_hdd_size": get_text("id_hdd_size"),
            "id_sys": get_text("id_sys"),
            "id_prn": get_text("id_prn"),
            "id_host": get_text("id_host"),
            "id_ip": get_text("id_ip"),
            "id_mac": get_text("id_mac"),
            "id_office": get_text("id_office"),
            "id_structure": get_text("id_structure"),
            "id_inv_number": get_text("id_inv_number"),
        }
        return data
    except Exception:
        return None

def is_file_locked(path: Path) -> bool:
    """
    Проверяет, заблокирован ли файл (например, открыт в Excel).
    Возвращает True, если файл нельзя открыть для записи.
    """
    try:
        with open(path, "ab") as f:
            return False
    except (PermissionError, OSError):
        return True

def ensure_html_with_marker(html_path: Path):
    """
    Если HTML-файла нет — создаёт его с базовым скелетом и маркером.
    Если файл есть, но маркера нет — добавляет маркер в конец.
    """
    if not html_path.exists():
        # Создаём минимальный HTML с маркером
        content = """<!DOCTYPE html>
<html>
  <head>
    <meta charset="utf-8">
    <title>List Points</title>
  </head>
  <body>
    <h1>Список точек</h1>
    <!-- INSERT_DIMON_BLOCK_HERE -->
  </body>
</html>
"""
        html_path.write_text(content, encoding="utf-8")
        return

    text = html_path.read_text(encoding="utf-8")
    if MARKER not in text:
        # Если маркера нет, добавляем его в конец перед закрывающим </body>
        if "</body>" in text:
            text = text.replace("</body>", f"{MARKER}\n</body>")
        else:
            text += f"\n{MARKER}\n"
        html_path.write_text(text, encoding="utf-8")

def append_html_table_rows(html_path: Path, rows_data: list[dict]):
    """
    Вставляет строки <tr>...</tr> сразу после маркера в HTML.
    rows_data — список словарей с данными.
    """
    text = html_path.read_text(encoding="utf-8")
    marker_pos = text.find(MARKER)
    if marker_pos == -1:
        raise RuntimeError(f"Маркер '{MARKER}' не найден в {html_path}")

    # Находим позицию сразу после маркера
    insert_pos = marker_pos + len(MARKER)

    # Формируем строки таблицы
    html_rows = []
    for row in rows_data:
        html_row = (
            f'\n      <tr>\n'
            f'        <td>{row.get("id_date", "")}</td>\n'
            f'        <td>{row.get("id_ip", "")}</td>\n'
            f'        <td>{row.get("id_mac", "")}</td>\n'
            f'        <td>{row.get("id_mb", "")}</td>\n'
            f'        <td>{row.get("id_cpu", "")}</td>\n'
            f'        <td>{row.get("id_mem", "")}</td>\n'
            f'        <td>{row.get("id_hdd", "")}</td>\n'
            f'        <td>{row.get("id_hdd_size", "")}</td>\n'
            f'        <td>{row.get("id_sys", "")}</td>\n'
            f'        <td>{row.get("id_host", "")}</td>\n'
            f'        <td>{row.get("id_prn", "")}</td>\n'
            f'        <td>{row.get("id_office", "")}</td>\n'
            f'        <td>{row.get("id_structure", "")}</td>\n'
            f'        <td>{row.get("id_inv_number", "")}</td>\n'
            f'      </tr>\n'
        )
        html_rows.append(html_row)

    new_text = text[:insert_pos] + "".join(html_rows) + text[insert_pos:]
    html_path.write_text(new_text, encoding="utf-8")

def main():
    xml_folder = Path(XML_FOLDER)
    xlsx_path = Path(XLSX_FILE)
    html_path = Path(HTML_FILE)

    # Проверка наличия папки и файлов
    if not xml_folder.exists():
        print(f"{Fore.RED}Папка '{xml_folder}' не найдена.{Style.RESET_ALL}")
        return

    xml_files = [f for f in xml_folder.iterdir() if f.suffix.lower() in (".xml", ".XML")]
    if not xml_files:
        print(f"{Fore.RED}В папке XML нет файлов .xml{Style.RESET_ALL}")
        return

    if not xlsx_path.exists():
        print(f"{Fore.RED}Файл '{xlsx_path}' не найден.{Style.RESET_ALL}")
        return

    # ПРОВЕРКА: если Excel-файл открыт — выходим с красным сообщением
    if is_file_locked(xlsx_path):
        print(
            f"{Fore.RED}"
            f"Ошибка: файл '{xlsx_path}' заблокирован (скорее всего, открыт в Excel).\n"
            f"Закройте его и запустите скрипт снова.{Style.RESET_ALL}"
        )
        return

    # Подготовка HTML-файла (создать/добавить маркер)
    ensure_html_with_marker(html_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active  # берём активный лист

    # Порядок колонок (должен соответствовать порядку столбцов в Excel)
    columns_order = [
        "id_date",
        "id_ip",
        "id_mac",
        "id_mb",
        "id_cpu",
        "id_mem",
        "id_hdd",
        "id_hdd_size",
        "id_sys",
        "id_host",
        "id_prn",
        "id_office",
        "id_structure",
        "id_inv_number",
    ]

    # Стили шрифтов
    calibri_12_normal = Font(name="Calibri", size=12, bold=False)
    calibri_12_bold_red    = Font(name="Calibri", size=12, bold=True, color="FF0000")   # 2-й столбец: красный
    calibri_12_bold_orange = Font(name="Calibri", size=12, bold=True, color="FF8C00")   # 3-й столбец: оранжевый

    # Границы: левая, правая, верхняя — белые; нижняя — серая
    white_side = Side(border_style="thin", color="FFFFFF")
    gray_bottom = Side(border_style="thin", color="808080")

    border_all_white_top_left_right = Border(
        left=white_side,
        right=white_side,
        top=white_side,
        bottom=gray_bottom
    )

    # Базовые выравнивания
    align_center_v = Alignment(vertical="center")

    success_count = 0
    fail_count = 0
    failed_files = []  # список необработанных файлов
    success_rows_for_html = []  # данные для HTML-таблицы

    # Прогресс-бар
    for xml_file in tqdm(xml_files, desc="Обработка XML"):
        data = parse_xml(xml_file)
        if data is None:
            fail_count += 1
            failed_files.append(xml_file.name)
            continue

        # Вставляем строку на позицию START_ROW (сдвиг вниз)
        ws.insert_rows(START_ROW)

        # Применяем высоту строки 18
        ws.row_dimensions[START_ROW].height = 18

        # Заполняем вставленную строку и применяем стили
        for col_idx, key in enumerate(columns_order, start=1):
            cell = ws.cell(row=START_ROW, column=col_idx, value=data[key])

            # Вертикальное выравнивание везде по центру
            cell.alignment = align_center_v

            # Горизонтальное выравнивание и шрифт по номеру столбца
            if col_idx == 1:
                # 1-й столбец: по центру, обычный шрифт
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.font = calibri_12_normal

            elif col_idx == 2:
                # 2-й столбец: IP — жирный красный, по умолчанию left
                cell.alignment = Alignment(vertical="center", horizontal="left")
                cell.font = calibri_12_bold_red

            elif col_idx == 3:
                # 3-й столбец: MAC — жирный оранжевый, по умолчанию left
                cell.alignment = Alignment(vertical="center", horizontal="left")
                cell.font = calibri_12_bold_orange

            else:
                # с 4-го столбца: по правому краю, обычный шрифт
                cell.alignment = Alignment(vertical="center", horizontal="right")
                cell.font = calibri_12_normal

            # Применяем границы ко всем ячейкам
            cell.border = border_all_white_top_left_right

        success_count += 1
        success_rows_for_html.append(data)

    wb.save(xlsx_path)

    # Записываем строки в HTML
    if success_rows_for_html:
        append_html_table_rows(html_path, success_rows_for_html)

    # Вывод статистики
    print("\n--- Статистика ---")
    print(f"Всего файлов: {len(xml_files)}")
    print(f"Успешно обработано: {success_count}")
    print(f"Не удалось обработать: {fail_count}")

    if failed_files:
        print("\nНеобработанные файлы:")
        for name in failed_files:
            print(f"  - {name}")

if __name__ == "__main__":
    main()
